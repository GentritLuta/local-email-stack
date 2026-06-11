"""build-curated-list.py — generate a curated metro referral list (clean CSV +
styled XLSX) from a verified metro JSON, and register it in curated.json. Same
format/style as the Indianapolis list. No fabrication: every field comes from the
metro JSON (built by research agents, primary-source verified).

Metro JSON shape (referral-lists/metros/<slug>.json):
  {"metro": "...", "label": "...", "basename": "Attorney-Referral-List-...",
   "match": {"area_codes": [...], "cities": [...]},
   "firms": [{"firm","type","lead_attorney","practice","city","phone","email",
              "website","address"}, ...]}    # type = "Estate / Probate" | "Divorce / Family"

Usage:
  py scripts/build-curated-list.py referral-lists/metros/columbus.json
"""
import csv
import html
import json
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
LISTS = REPO / "referral-lists"
CURATED = LISTS / "curated.json"
COLS = ["Firm", "Type", "Lead Attorney", "Practice Focus", "City",
        "Phone", "Email", "Website", "Address"]


def norm_phone(p: str) -> str:
    d = re.sub(r"\D", "", p or "")
    if len(d) == 11 and d.startswith("1"):
        d = d[1:]
    return f"{d[0:3]}-{d[3:6]}-{d[6:10]}" if len(d) == 10 else (p or "")


def to_rows(firms: list[dict]) -> list[dict]:
    rows = []
    for f in firms:
        rows.append({
            "Firm": html.unescape(f.get("firm", "")).strip(),
            "Type": f.get("type", "").strip(),
            "Lead Attorney": html.unescape(f.get("lead_attorney", "")).strip(),
            "Practice Focus": html.unescape(f.get("practice", "")).strip(),
            "City": f.get("city", "").strip(),
            "Phone": norm_phone(f.get("phone", "")),
            "Email": f.get("email", "").strip(),
            "Website": f.get("website", "").strip(),
            "Address": html.unescape(f.get("address", "")).strip(),
        })
    # Estate first, then Divorce; within a type, by city
    rows.sort(key=lambda r: (0 if r["Type"].startswith("Estate") else 1, r["City"]))
    return rows


def write_csv(rows, path):
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLS)
        w.writeheader(); w.writerows(rows)


def write_xlsx(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Attorney Referrals"
    head_fill = PatternFill("solid", fgColor="0F172A")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(bottom=Side(style="thin", color="E2E8F0"))
    ws.append(COLS)
    for ci in range(1, len(COLS) + 1):
        c = ws.cell(row=1, column=ci); c.fill = head_fill; c.font = head_font
        c.alignment = Alignment(vertical="center")
    blue = PatternFill("solid", fgColor="DBEAFE"); green = PatternFill("solid", fgColor="DCFCE7")
    blue_f = Font(color="1E40AF", bold=True, size=10); green_f = Font(color="166534", bold=True, size=10)
    for r in rows:
        ws.append([r[c] for c in COLS]); ri = ws.max_row
        is_div = r["Type"].startswith("Divorce")
        for ci in range(1, len(COLS) + 1):
            cell = ws.cell(row=ri, column=ci); cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(COLS[ci - 1] in ("Practice Focus", "Address")))
            if COLS[ci - 1] == "Type":
                cell.fill = blue if is_div else green
                cell.font = blue_f if is_div else green_f
                cell.alignment = Alignment(vertical="center", horizontal="center")
    widths = {"Firm": 34, "Type": 16, "Lead Attorney": 22, "Practice Focus": 34,
              "City": 16, "Phone": 14, "Email": 30, "Website": 24, "Address": 40}
    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 22
    wb.save(path)


def register(entry: dict):
    cur = json.loads(CURATED.read_text(encoding="utf-8"))
    lists = cur.setdefault("lists", [])
    lists[:] = [e for e in lists if e.get("metro") != entry["metro"]]
    lists.append(entry)
    CURATED.write_text(json.dumps(cur, indent=2), encoding="utf-8")


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: py scripts/build-curated-list.py <metro.json>")
    meta = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    rows = to_rows(meta["firms"])
    base = meta["basename"]
    csv_name, xlsx_name = base + ".csv", base + ".xlsx"
    write_csv(rows, LISTS / csv_name)
    write_xlsx(rows, LISTS / xlsx_name)
    register({"metro": meta["metro"], "label": meta["label"],
              "csv": csv_name, "xlsx": xlsx_name, "match": meta["match"]})
    n_e = sum(1 for r in rows if r["Type"].startswith("Estate"))
    n_mail = sum(1 for r in rows if r["Email"])
    print(f"{meta['metro']}: {len(rows)} firms ({n_e} estate, {len(rows)-n_e} divorce), "
          f"{n_mail} emails -> {csv_name} + {xlsx_name}; registered in curated.json")


if __name__ == "__main__":
    sys.exit(main())
