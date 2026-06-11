"""build-jake-list.py — single source of truth for Jake's attorney referral
list. 30 firms, each field VERIFIED from the firm's own website / reputable
directories by research agents (phones corrected vs directories, real published
emails only, full addresses, lead attorney, accurate practice focus).

Ordered: Estate/Probate first (Greenwood -> Indianapolis -> suburbs), then
Divorce/Family (Greenwood -> Indianapolis -> Carmel). Writes:

  referral-lists/Attorney-Referral-List-Indianapolis.xlsx   (styled, opens clean)
  referral-lists/Attorney-Referral-List-Indianapolis.csv    (UTF-8 BOM, columns)

build() returns the row dicts.
"""
from pathlib import Path
import csv

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "referral-lists"
XLSX = OUT / "Attorney-Referral-List-Indianapolis.xlsx"
CSVF = OUT / "Attorney-Referral-List-Indianapolis.csv"

COLS = ["Firm", "Type", "Lead Attorney", "Practice Focus", "City",
        "Phone", "Email", "Website", "Address"]

E = "Estate / Probate"
D = "Divorce / Family"

# (firm, type, lead_attorney, practice, city, phone, email, website, address)
FIRMS = [
    # ===== Estate / Probate — Greenwood (Jake's backyard) =====
    ("H.G. Myers Law", E, "Heather L. George Myers", "Estate planning, wills, trusts, probate", "Greenwood, IN", "317-643-5496", "", "hgmyerslaw.com", "107 N State Rd 135, Suite 204, Greenwood, IN 46142"),
    ("Walterman Legal", E, "Joseph B. Walterman", "Estate planning, family law, civil matters", "Greenwood, IN", "317-953-2281", "", "waltermanlegal.com", "200 S Madison Ave, Greenwood, IN 46142"),
    ("Hocker & Associates LLC", E, "Janet Davis Hocker", "Real estate, estate planning, probate", "Greenwood, IN", "317-215-5045", "", "hockerlawfirm.com", "3209 Smith Valley Rd, Suite 233, Greenwood, IN 46142"),
    ("Vick Law, P.C.", E, "Thomas A. Vick", "Probate, estate administration, wills and trusts", "Greenwood, IN", "317-593-9853", "support@vicklaw.org", "vicklaw.org", "3209 W Smith Valley Rd, Suite 113, Greenwood, IN 46142"),
    ("Henn Haworth Cummings + Page", E, "David M. Henn", "Estate planning, probate, real estate", "Greenwood, IN", "317-885-0041", "info@hhcplaw.com", "hhcplaw.com", "1634 W Smith Valley Rd, Suite B, Greenwood, IN 46142"),
    ("Van Valer Law Firm, LLP", E, "Tom Vander Luitgaren", "Real estate, probate, estate planning", "Greenwood, IN", "317-881-7575", "info@vanvalerlaw.com", "vanvalerlaw.com", "225 S Emerson Ave, Suite 181, Greenwood, IN 46143"),
    ("Law Office of Emily Greeson", E, "Emily Greeson", "Estate planning, elder law, probate", "Greenwood, IN", "317-385-3126", "emilygreeson.law@gmail.com", "emilygreesonelderlaw.com", "432 S Emerson Ave, Suite 330, Greenwood, IN 46143"),
    # ===== Estate / Probate — Indianapolis & suburbs =====
    ("Law Office of Melissa Winkler-York", E, "Melissa Winkler-York", "Elder law, wills, estate planning", "Indianapolis, IN", "317-781-1080", "info@winkleryorklaw.com", "winkleryorklaw.com", "4259 Shelby St, Indianapolis, IN 46227"),
    ("Frank & Kraft, Attorneys at Law", E, "Paul Kraft", "Estate planning, probate, trust administration", "Indianapolis, IN", "317-684-1100", "paul@frankkraft.com", "frankkraft.com", "135 N Pennsylvania St, Suite 1100, Indianapolis, IN 46204"),
    ("Marc Matheny Law", E, "Marc W. Matheny", "Probate, estate administration, wills and trusts", "Indianapolis, IN", "317-639-3315", "info@marcmathenylaw.com", "marcmathenylaw.com", "244 N College Ave, Indianapolis, IN 46202"),
    ("Harshman Ponist Smith & Rayl, LLC", E, "Aaron J. Harshman", "Estate planning, probate, civil litigation", "Indianapolis, IN", "317-964-6000", "", "hpindiana.law", "3650 N Washington Blvd, Indianapolis, IN 46205"),
    ("Robert W. York & Associates", E, "Robert W. York", "Estate planning, probate, probate litigation", "Indianapolis, IN", "317-842-8000", "", "york-law.com", "7212 N Shadeland Ave, Suite 150, Indianapolis, IN 46250"),
    ("Barnes Cadwell Law, P.A.", E, "Martin Barnes", "Estate planning, probate, asset protection", "Indianapolis, IN", "317-804-5058", "", "barnescadwell.com", "10475 Crosspoint Blvd, Suite 250, Indianapolis, IN 46256"),
    ("Adler Law LLC", E, "Michael J. Adler", "Estate planning, wills, trusts", "Indianapolis, IN", "317-635-7880", "", "adlerlawonline.com", "727 E 86th St, Indianapolis, IN 46240"),
    ("Indy Advocate (McNevin & McInnes)", E, "Rob McNevin", "Real estate, probate, wills and trusts", "Indianapolis, IN", "317-939-3000", "", "indyadvocate.com", "5224 S East St, Suite C14, Indianapolis, IN 46227"),
    ("Allen Wellman Harvey Keyes Cooley, LLP", E, "Eric N. Allen", "Estate planning, probate, real estate", "Greenfield, IN", "317-468-2355", "", "awmlaw.com", "Five Courthouse Plaza, Greenfield, IN 46140"),
    ("Holwager & Holwager", E, "Bill Holwager", "Elder law, estate planning, probate", "Beech Grove, IN", "317-787-8395", "intake@hhelderlaw.com", "hhelderlaw.com", "1818 Main St, Beech Grove, IN 46107"),
    # ===== Divorce / Family =====
    ("Howe Law", D, "Martin N. Howe", "Family law, divorce, probate", "Greenwood, IN", "317-644-2557", "", "howefamilylaw.com", "386 Meridian Parke Lane, Suite A, Greenwood, IN 46142"),
    ("The Renner Law Office", D, "Stephanie Renner", "Divorce, family law, prenuptial agreements", "Indianapolis, IN", "317-771-8535", "Stephanie@TheRennerLawOffice.com", "therennerlawoffice.com", "1512 N Delaware St, Indianapolis, IN 46202"),
    ("Lopez Law Office", D, "Vanessa Lopez", "Divorce, custody, child support", "Indianapolis, IN", "317-634-9484", "", "vanessalopezlaw.com", "3502 N Meridian St, Indianapolis, IN 46208"),
    ("Stange Law Firm, PC", D, "Kirk Stange", "Divorce, property division, estate planning", "Indianapolis, IN", "463-258-5401", "", "stangelawfirm.com", "3905 Vincennes Rd, Suite 103, Indianapolis, IN 46268"),
    ("Eskew Law, LLC", D, "Christopher Eskew", "Family law, divorce, criminal defense", "Indianapolis, IN", "317-974-0177", "", "eskewlaw.com", "255 N Alabama St, 3rd Floor, Indianapolis, IN 46204"),
    ("Keffer Hirschauer LLP", D, "Bradley Keffer", "Family law, divorce, expungement", "Indianapolis, IN", "317-857-0160", "help@indyjustice.com", "indyjustice.com", "230 E Ohio St, Suite 400, Indianapolis, IN 46204"),
    ("Villarrubia & Rosenberger, PC", D, "Tabitha Villarrubia", "Family law, divorce, custody", "Indianapolis, IN", "463-207-9900", "", "vrlegal.com", "6349 S East St, Indianapolis, IN 46227"),
    ("Doyle Family Law LLC", D, "Eric K. Doyle", "Divorce, child custody, support", "Indianapolis, IN", "317-505-0971", "", "doylefamilylawyers.com", "1300 E 86th St, #36A, Indianapolis, IN 46240"),
    ("Cordell & Cordell", D, "Robert Haywood", "Divorce, custody, marital property division", "Indianapolis, IN", "317-322-0122", "", "cordellcordell.com", "101 W Ohio St, Suite 1250, Indianapolis, IN 46204"),
    ("Emerson Divorce Attorneys", D, "JR Emerson", "Divorce, custody, high-asset property division", "Carmel, IN", "317-969-8000", "", "emersonlawllc.com", "1 S Rangeline Rd, Suite 400, Carmel, IN 46032"),
    ("Coots, Henke & Wheeler", D, "Daniel E. Coots", "Family law, probate, estate planning", "Carmel, IN", "317-708-4819", "", "chwlaw.com", "255 E Carmel Dr, Carmel, IN 46032"),
    ("Cross Glazier Reed Burroughs, PC", D, "Nancy L. Cross", "Divorce, custody, marital property division", "Carmel, IN", "317-582-2053", "", "cgblawfirm.com", "11595 N Meridian St, Suite 110, Carmel, IN 46032"),
    ("Connell Michael LLP", D, "", "Divorce, custody, family law", "Carmel, IN", "317-434-1061", "", "connellmichaellaw.com", "550 Congressional Blvd, Suite 350-11, Carmel, IN 46032"),
]


def rows() -> list[dict]:
    return [dict(zip(COLS, f)) for f in FIRMS]


def _write_csv(data: list[dict]) -> None:
    with CSVF.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLS)
        w.writeheader()
        w.writerows(data)


def _write_xlsx(data: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Attorney Referrals"

    head_fill = PatternFill("solid", fgColor="0F172A")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(bottom=Side(style="thin", color="E2E8F0"))

    ws.append(COLS)
    for ci in range(1, len(COLS) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = head_fill; c.font = head_font
        c.alignment = Alignment(vertical="center")

    blue = PatternFill("solid", fgColor="DBEAFE")
    green = PatternFill("solid", fgColor="DCFCE7")
    blue_f = Font(color="1E40AF", bold=True, size=10)
    green_f = Font(color="166534", bold=True, size=10)
    for r in data:
        ws.append([r[c] for c in COLS])
        ri = ws.max_row
        is_div = r["Type"].startswith("Divorce")
        for ci in range(1, len(COLS) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(COLS[ci - 1] in ("Practice Focus", "Address")))
            if COLS[ci - 1] == "Type":
                cell.fill = blue if is_div else green
                cell.font = blue_f if is_div else green_f
                cell.alignment = Alignment(vertical="center", horizontal="center")

    widths = {"Firm": 34, "Type": 16, "Lead Attorney": 22, "Practice Focus": 34,
              "City": 16, "Phone": 14, "Email": 30, "Website": 24, "Address": 40}
    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(data) + 1}"
    ws.row_dimensions[1].height = 22
    wb.save(XLSX)


def build() -> tuple[list[dict], Path, Path]:
    OUT.mkdir(exist_ok=True)
    data = rows()
    _write_csv(data)
    _write_xlsx(data)
    return data, XLSX, CSVF


if __name__ == "__main__":
    data, x, c = build()
    n_e = sum(1 for r in data if r["Type"].startswith("Estate"))
    n_mail = sum(1 for r in data if r["Email"])
    n_phone = sum(1 for r in data if r["Phone"])
    print(f"wrote {len(data)} firms: {n_e} estate/probate, {len(data)-n_e} divorce/family")
    print(f"  phones filled: {n_phone}/{len(data)} | emails: {n_mail}/{len(data)}")
    print(" ", x)
    print(" ", c)
